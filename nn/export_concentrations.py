#!/usr/bin/env python3
"""
Export initial and feed concentrations (current vs proposed) to Excel.

Proposed values:
  - All amino acids (indices 6-24): 210x DMEM base, applied to both initial and feed
  - NH4 feed: 1.1 mmol/L (initial unchanged at 0)
  - Glucose feed: 25.0 mmol/L (initial unchanged at 17.5)

Usage:
    python export_concentrations.py
"""

import numpy as np
import pandas as pd
from pathlib import Path

from generate_synthetic_ode import C_NOMINAL, CIN_NOMINAL, AAS_INDICES

AA_SCALE  = 210.0
NH4_FEED  = 1.1
GLC_FEED  = 25.0

COMPONENT_NAMES = [
    'Cell Density', 'Cell Volume', 'Glucose', 'Lactate', 'NH4', 'Titer',
    'Glutamine', 'Glutamate', 'L-Asparagine', 'L-Aspartic acid',
    'L-Serine', 'Glycine', 'L-Alanine', 'L-Proline', 'L-Threonine',
    'L-Histidine', 'L-Lysine', 'L-Valine', 'L-Methionine', 'L-Arginine',
    'L-Tyrosine', 'L-Isoleucine', 'L-Leucine', 'L-Phenylalanine',
    'L-Tryptophan',
]

UNITS = (
    ['E9 cells/L', 'um3/1000', 'mmol/L', 'mmol/L', 'mmol/L', 'mg/L'] +
    ['mmol/L'] * 19
)

# Build proposed arrays
c_nom_new = C_NOMINAL.copy()
cin_nom_new = CIN_NOMINAL.copy()

c_nom_new[AAS_INDICES]   *= AA_SCALE
cin_nom_new[AAS_INDICES] *= AA_SCALE
cin_nom_new[4]            = NH4_FEED   # NH4 feed only
cin_nom_new[2]            = GLC_FEED   # Glucose feed only

rows = []
for i, (name, unit) in enumerate(zip(COMPONENT_NAMES, UNITS)):
    rows.append({
        'Component':            name,
        'Units':                unit,
        'Initial (current)':    C_NOMINAL[i],
        'Initial (proposed)':   c_nom_new[i],
        'Feed (current)':       CIN_NOMINAL[i],
        'Feed (proposed)':      cin_nom_new[i],
    })

df = pd.DataFrame(rows)

out = Path(__file__).parent / 'proposed_concentrations.xlsx'
with pd.ExcelWriter(out, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Concentrations')

    ws = writer.sheets['Concentrations']

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18

    # Highlight changed rows
    from openpyxl.styles import PatternFill
    highlight = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    for row_idx, i in enumerate(range(len(df)), start=2):
        changed = (
            c_nom_new[i] != C_NOMINAL[i] or
            cin_nom_new[i] != CIN_NOMINAL[i]
        )
        if changed:
            for col_idx in range(1, 7):
                ws.cell(row=row_idx, column=col_idx).fill = highlight

print(f'Saved: {out}')
print(f'\nSummary of changes:')
print(f'  AA scale (indices 6-24): {AA_SCALE}x applied to initial and feed')
print(f'  NH4 feed: {CIN_NOMINAL[4]:.3f} -> {NH4_FEED} mmol/L')
print(f'  Glucose feed: {CIN_NOMINAL[2]:.3f} -> {GLC_FEED} mmol/L')
